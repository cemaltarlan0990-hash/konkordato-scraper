"""
Konkordato ilan <-> CRM firma eslestirme motoru.

Tek kural seti, iki tarafa da AYNI sekilde uygulanir.
CRM tarafi ve ilan tarafi ayni fonksiyondan gecer -> asimetri imkansiz.
"""

import json
import csv
import unicodedata


# ---------------------------------------------------------------------------
# AYARLAR
# ---------------------------------------------------------------------------

# Turkce/Latin karakter katlamasi (BIOSCIENCE == BIOSCIENCE)
# Kapatmak icin False yap. Iki tarafa birden uygulanir.
KATLAMA = True

# Eslestirmede hangi isim sutunu kullanilacak:
#   "orijinal"    - sadece OrijinalIsim (CRM'deki ham hali)
#   "duzenlenmis" - sadece DuzenlenmisFirmaAdi (elle duzeltilmis hali)
#   "her_ikisi"   - ikisi de denenir, biri tutarsa eslesme sayilir
# Mail her durumda OrijinalIsim gosterir.
ISIM_KAYNAGI = "her_ikisi"

# Sondan silinecek unvan kelimeleri (ham hali; kod bunlari da ayni
# normalizasyondan gecirir, boylece katlama acik/kapali fark etmez)
UNVAN_KELIMELERI_HAM = [
    "AS", "A.S.", "ANONIM", "SIRKETI", "STI", "LTD", "LIMITED",
    "KOLLEKTIF", "KOMANDIT", "ADI", "KOOPERATIFI", "KOOPERATIF",
    "AŞ", "ANONİM", "ŞİRKETİ", "ŞTİ", "LİMİTED",
    "KOLLEKTİF", "KOMANDİT", "ADİ", "KOOPERATİFİ",
    "ORTAKLIGI", "ORTAKLIĞI", "SUBESI", "ŞUBESİ",
]

# Jenerik sektor/kuyruk kelimeleri.
# Sebep: CRM "SAN. VE TIC." yazarken ilan "SANAYI VE TICARET" yaziyor.
# Bunlar kelime kelime karsilastirmayi bozuyordu (ATABAY KIDS TEKSTIL
# tam tutmasina ragmen skor 3/6'ya dusuyordu).
#
# Bu bir sozluk/kanoniklestirme DEGIL - ayni "sondan silme" mekanizmasi,
# sadece daha genis liste. Ortadaki gecisler korunur; SADECE SONDAN silinir.
# Kapatmak icin JENERIK_KUYRUK_SIL = False.
JENERIK_KUYRUK_SIL = True

JENERIK_KELIMELER_HAM = [
    "SANAYI", "SANAYİ", "SANAYII", "SAN",
    "TICARET", "TİCARET", "TIC", "TİC",
    "VE", "ILE", "İLE",
    "PAZARLAMA", "ITHALAT", "İTHALAT", "IHRACAT", "İHRACAT",
    "DIS", "DIŞ", "IC", "İÇ",
]

# CRM referansinda beklenen minimum kayit sayisi.
# Bunun altina duserse veri bozuk demektir -> calisma durur.
MIN_CRM_KAYIT = 10000


# ---------------------------------------------------------------------------
# 1. TEMIZLIK VE NORMALIZASYON
# ---------------------------------------------------------------------------

def cift_nokta_temizle(metin):
    """
    Bozukluk: dogru olusmus I (U+0130) karakterinin hemen ardina
    fazladan bir COMBINING DOT ABOVE (U+0307) eklenmis.
    Gorsel olarak ayni, string olarak farkli.
    NFC normalizasyonu bunu YAKALAMAZ, acikca silmek gerekir.
    """
    if not metin:
        return ""
    metin = metin.replace("\u0130\u0307", "\u0130")
    metin = metin.replace("\u0049\u0307", "\u0130")
    return unicodedata.normalize("NFC", metin)


def turkce_upper(metin):
    """
    Turkce-farkindalikli buyuk harfe cevirme.
    SIRA KRITIK: once i->I ve i->I esleme, sonra upper().
    Ters sirada Python'un upper()'i i icin bilesen nokta uretiyor.
    """
    metin = metin.replace("i", "\u0130").replace("\u0131", "I")
    return metin.upper()


def katla(metin):
    """
    Turkce karakterleri Latin karsiliklarina indirger.
    Amac: CRM ve ilan taraflarindaki kodlama tutarsizligini anlamsiz kilmak.
    Sadece karsilastirma aninda kullanilir; hicbir dosyaya yazilmaz.
    """
    tablo = {
        "\u0130": "I", "\u0131": "I", "\u0049": "I", "\u0069": "I",
        "\u015E": "S", "\u015F": "S",
        "\u011E": "G", "\u011F": "G",
        "\u00C7": "C", "\u00E7": "C",
        "\u00D6": "O", "\u00F6": "O",
        "\u00DC": "U", "\u00FC": "U",
    }
    return "".join(tablo.get(k, k) for k in metin)


def kelimelere_ayir(metin):
    """
    Bosluk VE nokta ile ayir.
    Sadece boslukla ayirmak "SAN.VE TIC.LTD.STI." gibi bitisik
    kisaltmalari tek kelime birakiyordu.
    Diger noktalama isaretleri de ayirici sayilir.
    """
    for isaret in [".", ",", ";", ":", "/", "\\", "-", "(", ")", "\"", "'"]:
        metin = metin.replace(isaret, " ")
    return [k for k in metin.split() if k]


def tek_harfleri_birlestir(kelimeler):
    """
    Ardisik tek-harfli kelimeleri birlestirir.
    Nokta ile ayirma "A.S."yi "A" + "S" diye parcaliyordu.
    Bu adim olmadan sondan unvan silme calismaz.
    15.088 kaydin 3.581'ini (%23,7) etkileyen bir hataydi.
    """
    sonuc = []
    tampon = ""
    for kelime in kelimeler:
        if len(kelime) == 1:
            tampon += kelime
        else:
            if tampon:
                sonuc.append(tampon)
                tampon = ""
            sonuc.append(kelime)
    if tampon:
        sonuc.append(tampon)
    return sonuc


def _liste_normalize(ham_liste, katlama):
    """Kelime listesini ayni normalizasyondan gecirip set olarak dondurur."""
    hazir = set()
    for ham in ham_liste:
        temiz = turkce_upper(cift_nokta_temizle(ham))
        if katlama:
            temiz = katla(temiz)
        for parca in tek_harfleri_birlestir(kelimelere_ayir(temiz)):
            hazir.add(parca)
    return hazir


def _unvan_seti(katlama):
    """
    Sondan silinecek kelime setleri.
    Iki asamali: once hukuki unvanlar, sonra jenerik kuyruk.
    Doner: (unvan_seti, jenerik_seti)
    """
    unvanlar = _liste_normalize(UNVAN_KELIMELERI_HAM, katlama)
    jenerik = set()
    if JENERIK_KUYRUK_SIL:
        jenerik = _liste_normalize(JENERIK_KELIMELER_HAM, katlama)
    return unvanlar, jenerik


def sondan_unvan_sil(kelimeler, unvanlar):
    """
    Unvan kelimelerini SADECE SONDAN, art arda siler.
    Ortadaki SANAYI/TICARET firmanin gercek adinin parcasi olabilir,
    o yuzden ortaya dokunulmaz.
    """
    sonuc = list(kelimeler)
    while sonuc and sonuc[-1] in unvanlar:
        sonuc.pop()
    return sonuc


def saf_kelimeler(isim, katlama=KATLAMA, setler=None):
    """
    Bir firma adindan cekirdek kelime listesi uretir.
    CRM ve ilan taraflarina AYNI SEKILDE uygulanir.

    Silme iki asamali ve GERI ALINABILIR:
      1. Hukuki unvanlar sondan silinir
      2. Jenerik kuyruk sondan silinir
    Bir asama listeyi tamamen bosaltirsa o asama GERI ALINIR.
    Boylece "SANAYI VE TICARET A.S." gibi tamamen jenerik bir isim
    bos listeye dusmez - elde kalan neyse onunla devam edilir.
    """
    if setler is None:
        setler = _unvan_seti(katlama)
    unvanlar, jenerik = setler

    metin = cift_nokta_temizle(isim)
    metin = turkce_upper(metin)
    if katlama:
        metin = katla(metin)

    kelimeler = tek_harfleri_birlestir(kelimelere_ayir(metin))
    if not kelimeler:
        return []

    asama1 = sondan_unvan_sil(kelimeler, unvanlar)
    if not asama1:
        return kelimeler

    if not jenerik:
        return asama1

    asama2 = sondan_unvan_sil(asama1, jenerik)
    if not asama2:
        return asama1
    return asama2


# ---------------------------------------------------------------------------
# 2. CRM REFERANSI
# ---------------------------------------------------------------------------

class CRMReferans:
    """
    CRM firma listesini bellege alir ve her kayit icin saf kelimeleri
    CALISMA ANINDA uretir. Dosyadaki hazir SafKelimeler sutunu
    KULLANILMAZ - iki tarafin ayni koddan gecmesi garantisi icin.

    Iki isim sutunu desteklenir:
      OrijinalIsim        - CRM'deki ham hali. Mail'de BU gosterilir,
                            cunku personel CRM'de bu isimle arama yapiyor.
      DuzenlenmisFirmaAdi - elle duzeltilmis hali (yazim hatalari,
                            eksik/hatali karakterler). Eslestirmede
                            daha guvenilir.

    ISIM_KAYNAGI ile hangisinin eslestirmede kullanilacagi secilir.
    "her_ikisi" modunda her iki yazimin kelime dizisi de aday havuzuna
    girer; biri tutarsa eslesme sayilir. Mail her durumda OrijinalIsim
    gosterir.
    """

    def __init__(self, kayitlar, katlama=KATLAMA, isim_kaynagi=None):
        self.katlama = katlama
        self.isim_kaynagi = isim_kaynagi or ISIM_KAYNAGI
        self.setler = _unvan_seti(katlama)
        self.kayitlar = []
        self.vkn_index = {}

        for kayit in kayitlar:
            orijinal = (kayit.get("OrijinalIsim") or "").strip()
            duzenlenmis = (kayit.get("DuzenlenmisFirmaAdi") or "").strip()

            # Mail'de her zaman OrijinalIsim gosterilir.
            # Bos ise duzenlenmise duser.
            gosterilecek = orijinal or duzenlenmis
            if not gosterilecek:
                continue

            adaylar = []
            if self.isim_kaynagi in ("orijinal", "her_ikisi") and orijinal:
                adaylar.append(orijinal)
            if self.isim_kaynagi in ("duzenlenmis", "her_ikisi") and duzenlenmis:
                adaylar.append(duzenlenmis)
            if not adaylar:
                adaylar = [gosterilecek]

            kelime_dizileri = []
            for ad in adaylar:
                kelimeler = saf_kelimeler(ad, katlama, self.setler)
                if kelimeler and kelimeler not in kelime_dizileri:
                    kelime_dizileri.append(kelimeler)
            if not kelime_dizileri:
                continue

            vkn = str(kayit.get("VKN") or "").strip()
            satir = {
                "orijinal": gosterilecek,
                "vkn": vkn,
                # Her yazim varyanti ayri bir arama satiri olur;
                # hepsi ayni CRM kaydini gosterir.
                "kelimeler": kelime_dizileri[0],
                "tumKelimeler": kelime_dizileri,
            }
            self.kayitlar.append(satir)
            if vkn:
                self.vkn_index.setdefault(vkn, []).append(satir)

    def __len__(self):
        return len(self.kayitlar)

    def dogrula(self, minimum=MIN_CRM_KAYIT):
        """Bozuk/eksik referans dosyasiyla sessizce calismayi engeller."""
        if len(self.kayitlar) < minimum:
            raise ValueError(
                "CRM referansi beklenenden kucuk: %d kayit "
                "(minimum %d). Dosya eksik veya bozuk olabilir."
                % (len(self.kayitlar), minimum)
            )


# ---------------------------------------------------------------------------
# 3. ESLESTIRME
# ---------------------------------------------------------------------------

def vkn_ara(referans, vkn):
    """Vergi numarasi ile tam eslesme. Belirsizlik yok."""
    vkn = str(vkn or "").strip()
    if not vkn:
        return []
    return referans.vkn_index.get(vkn, [])


def kademeli_ara(referans, ilan_kelimeleri):
    """
    Soldan saga, kelime kelime aday havuzunu daraltir.
    Karsilastirma POZISYONELDIR: ilanin i. kelimesi, adayin i. kelimesiyle.

    Bir kelime hicbir adayda bulunamazsa DURUR - atlama yoktur.
    (Atlama denendi, 3 yanlis pozitif uretti.)

    Doner: (eslesen_kelime_sayisi, kalan_adaylar)
    """
    if not ilan_kelimeleri:
        return 0, []

    adaylar = referans.kayitlar
    eslesen = 0

    for i, kelime in enumerate(ilan_kelimeleri):
        yeni = [
            a for a in adaylar
            if any(len(kd) > i and kd[i] == kelime
                   for kd in a["tumKelimeler"])
        ]
        if not yeni:
            break
        adaylar = yeni
        eslesen = i + 1

    return eslesen, adaylar


def unvan_eslestir(referans, ilan_id, unvan, vkn=None):
    """
    Tek bir unvani eslestirir.
    Once VKN denenir (kesin), yoksa isim uzerinden kademeli arama.
    """
    sonuc = {
        "ilanId": ilan_id,
        "unvan": unvan,
        "vkn": vkn or "",
        "yontem": None,
        "durum": "NO MATCH",
        "skor": 0.0,
        "eslesenKelime": 0,
        "toplamKelime": 0,
        "adaylar": [],
    }

    # --- VKN yolu ---
    if vkn:
        bulunan = vkn_ara(referans, vkn)
        if bulunan:
            sonuc["yontem"] = "VKN"
            sonuc["durum"] = "MATCH"
            sonuc["skor"] = 1.0
            sonuc["adaylar"] = [
                {"crmIsim": b["orijinal"], "crmVkn": b["vkn"]} for b in bulunan
            ]
            return sonuc

    # --- Isim yolu ---
    kelimeler = saf_kelimeler(unvan, referans.katlama, referans.setler)
    sonuc["yontem"] = "ISIM"
    sonuc["toplamKelime"] = len(kelimeler)

    if not kelimeler:
        return sonuc

    # Cekirdek tamamen jenerik kelimelerden olusuyorsa arama yapma.
    # Gercek ornek: scraper "Ticaret A.S" diye kesik bir unvan uretti,
    # cekirdek sadece TICARET kaldi ve 5 alakasiz Bakanlik kaydiyla
    # eslesti. Ayirt edici tek kelime yoksa sonuc anlamsizdir.
    _, jenerik = referans.setler
    if jenerik and all(k in jenerik for k in kelimeler):
        sonuc["durum"] = "NO MATCH"
        sonuc["not"] = "Cekirdek tamamen jenerik, arama yapilmadi"
        return sonuc

    eslesen, adaylar = kademeli_ara(referans, kelimeler)
    sonuc["eslesenKelime"] = eslesen
    sonuc["skor"] = round(eslesen / len(kelimeler), 3)

    if eslesen == 0:
        return sonuc

    sonuc["adaylar"] = [
        {
            "crmIsim": a["orijinal"],
            "crmVkn": a["vkn"],
            "crmKelimeSayisi": max(len(kd) for kd in a["tumKelimeler"]),
        }
        for a in adaylar[:10]
    ]

    tam = (eslesen == len(kelimeler))
    if tam and len(adaylar) == 1:
        sonuc["durum"] = "MATCH"
    elif tam:
        # Ayni cekirdek ada sahip birden fazla CRM kaydi (cakisma grubu)
        sonuc["durum"] = "REVIEW"
    else:
        sonuc["durum"] = "REVIEW" if eslesen >= 2 else "NO MATCH"

    return sonuc


def ilanlari_isle(referans, ilanlar):
    """
    Tum ilanlari isler.

    VKN-unvan eslemesi SADECE firmalar[] uzerinden kurulur.
    unvanlar[] ile vergi numaralari arasinda pozisyonel esleme YOKTUR -
    scraper iki listeyi farkli siralarda dolduruyor.
    """
    sonuclar = []

    for ilan in ilanlar:
        ilan_id = ilan.get("ilanId") or ilan.get("id")
        # DIKKAT: ilanin "durum" alani (Gecici Muhlet vs.) eslesme
        # durumuyla (MATCH/REVIEW) CAKISIR. Ayri isimle tasiniyor.
        ek_alanlar = {
            alan: ilan[alan]
            for alan in ("tarih", "sehir", "mahkeme", "esas", "link")
            if alan in ilan
        }
        if "durum" in ilan:
            ek_alanlar["ilanDurumu"] = ilan["durum"]

        islenen_unvanlar = set()

        # --- 1. Unvani ve VKN'si birlikte yakalanmis firmalar ---
        for firma in (ilan.get("firmalar") or []):
            unvan = (firma.get("firma") or "").strip()
            vkn = (firma.get("vergiNo") or "").strip()
            if not unvan and not vkn:
                continue
            sonuc = unvan_eslestir(referans, ilan_id, unvan, vkn)
            sonuc.update(ek_alanlar)
            sonuclar.append(sonuc)
            if unvan:
                islenen_unvanlar.add(unvan)

        # --- 2. VKN'siz unvanlar (isim yoluyla) ---
        for unvan in (ilan.get("unvanlar") or []):
            unvan = str(unvan or "").strip()
            if not unvan or unvan in islenen_unvanlar:
                continue
            sonuc = unvan_eslestir(referans, ilan_id, unvan, None)
            sonuc.update(ek_alanlar)
            sonuclar.append(sonuc)
            islenen_unvanlar.add(unvan)

        # --- 3. Unvana baglanamamis serbest VKN'ler ---
        for vkn in (ilan.get("serbestVergiNolari") or []):
            bulunan = vkn_ara(referans, str(vkn).strip())
            if not bulunan:
                continue
            sonuc = {
                "ilanId": ilan_id,
                "unvan": "",
                "vkn": str(vkn).strip(),
                "yontem": "VKN",
                "durum": "MATCH",
                "skor": 1.0,
                "eslesenKelime": 0,
                "toplamKelime": 0,
                "adaylar": [
                    {"crmIsim": b["orijinal"], "crmVkn": b["vkn"]}
                    for b in bulunan
                ],
            }
            sonuc.update(ek_alanlar)
            sonuclar.append(sonuc)

    return sonuclar


# ---------------------------------------------------------------------------
# 4. DOSYA OKUMA
# ---------------------------------------------------------------------------

def crm_oku_csv(yol):
    """CSV referans dosyasini okur. Sutunlar: OrijinalIsim, VKN"""
    kayitlar = []
    with open(yol, encoding="utf-8-sig", newline="") as f:
        ornek = f.read(4096)
        f.seek(0)
        try:
            ayirici = csv.Sniffer().sniff(ornek, delimiters=",;\t").delimiter
        except Exception:
            ayirici = ","
        for satir in csv.DictReader(f, delimiter=ayirici):
            kayitlar.append(satir)
    return kayitlar


def crm_oku_excel(yol, sayfa=None):
    """XLSX referans dosyasini okur (openpyxl gerekir)."""
    from openpyxl import load_workbook
    wb = load_workbook(yol, read_only=True, data_only=True)
    ws = wb[sayfa] if sayfa else wb.active
    satirlar = ws.iter_rows(values_only=True)
    basliklar = [str(b).strip() if b else "" for b in next(satirlar)]
    kayitlar = []
    for satir in satirlar:
        kayitlar.append({
            basliklar[i]: satir[i]
            for i in range(min(len(basliklar), len(satir)))
        })
    wb.close()
    return kayitlar


def crm_oku(yol, sayfa=None):
    if yol.lower().endswith((".xlsx", ".xlsm")):
        return crm_oku_excel(yol, sayfa)
    return crm_oku_csv(yol)


# ---------------------------------------------------------------------------
# 5. CIKTI
# ---------------------------------------------------------------------------

def cikti_uret(sonuclar, uretim_zamani=None, crm_kayit_sayisi=0):
    """
    eslesmeler.json yapisi.
    uretimZamani PA tarafinda tazelik kontrolu icin kullanilir:
    dosya bugune ait degilse Actions patlamis demektir.
    """
    from datetime import datetime, timezone
    if uretim_zamani is None:
        uretim_zamani = datetime.now(timezone.utc).isoformat()

    return {
        "uretimZamani": uretim_zamani,
        "crmKayitSayisi": crm_kayit_sayisi,
        "katlama": KATLAMA,
        "toplamUnvan": len(sonuclar),
        "ozet": {
            "match": sum(1 for s in sonuclar if s["durum"] == "MATCH"),
            "review": sum(1 for s in sonuclar if s["durum"] == "REVIEW"),
            "noMatch": sum(1 for s in sonuclar if s["durum"] == "NO MATCH"),
        },
        "vknEslesmeleri": [
            s for s in sonuclar if s["yontem"] == "VKN" and s["durum"] == "MATCH"
        ],
        "isimEslesmeleri": [
            s for s in sonuclar if s["yontem"] == "ISIM" and s["durum"] == "MATCH"
        ],
        "incelenecekler": [s for s in sonuclar if s["durum"] == "REVIEW"],
    }


def calistir(crm_yolu, ilanlar, cikti_yolu="eslesmeler.json", sayfa=None):
    """Uctan uca calistirma."""
    referans = CRMReferans(crm_oku(crm_yolu, sayfa))
    referans.dogrula()
    sonuclar = ilanlari_isle(referans, ilanlar)
    cikti = cikti_uret(sonuclar, crm_kayit_sayisi=len(referans))
    with open(cikti_yolu, "w", encoding="utf-8") as f:
        json.dump(cikti, f, ensure_ascii=False, indent=2)
    return cikti
